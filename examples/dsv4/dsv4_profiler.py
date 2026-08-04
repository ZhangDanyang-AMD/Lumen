"""DSV4 Megatron torch.profiler hook and operator-breakdown xlsx export.

Env (rank 0 only, off unless ``LUMEN_PROF_START`` is set):

- ``LUMEN_PROF_START`` / ``LUMEN_PROF_END`` (default start+2): profile window
- ``LUMEN_PROF_OUTPUT``: text summary path (default under examples/dsv4/results/)
- ``LUMEN_PROF_XLSX``: operator breakdown xlsx path
- ``LUMEN_PROF_TRACE``: optional chrome trace path
- ``LUMEN_PROF_SHAPES``: record input shapes (0/1)
- ``LUMEN_PROF_STOP_AFTER``: exit training after this step (skip eval)
"""

from __future__ import annotations

import json
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from typing import Any


@dataclass(frozen=True)
class OpStat:
    category: str
    name: str
    count: int
    self_cuda_us: float
    cuda_us: float = 0.0
    input_shapes: str = ""

    @property
    def self_cuda_ms(self) -> float:
        return self.self_cuda_us / 1000.0


def _default_results_dir() -> str:
    here = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(here, "results")


def categorize_operator(name: str) -> str:
    n = name.lower()
    if "nccl" in n or "rccl" in n or "all_gather" in n or "reduce_scatter" in n:
        return "通信 (NCCL/RCCL/MoE EP)"
    if "gemm" in n or "mm_" in n or "baddbmm" in n or "bmm" in n:
        if any(k in n for k in ("a8w8", "fp8", "blockscale", "e4m3", "e5m2")):
            return "FP8 GEMM"
        return "GEMM/BMM"
    if any(k in n for k in ("mla", "fmha", "flash", "attention", "sparse_mla")):
        return "Sparse MLA / Attention"
    if any(k in n for k in ("mhc", "sinkhorn", "hyper_conn")):
        return "MHC (TileLang)"
    if any(k in n for k in ("indexer", "dsa_index", "lightning_indexer")):
        return "DSA Indexer"
    if any(k in n for k in ("moe", "expert", "dispatch", "combine", "alltoall", "router")):
        return "MoE router/dispatch"
    if any(k in n for k in ("copy_", "clone", "contiguous", "cat", "split")):
        return "拷贝/cat"
    if any(k in n for k in ("elementwise", "vectorized", "dropout", "swiglu", "silu")):
        return "elementwise"
    if "norm" in n or "rms" in n:
        return "Norm"
    if "tilelang" in n or n.startswith("tl_"):
        return "TileLang (other)"
    return "其他"


def kernel_type(category: str, name: str) -> str:
    n = name.lower()
    if "通信" in category or "nccl" in n or "rccl" in n:
        return "comm"
    if any(k in category for k in ("GEMM", "MLA", "Attention", "MHC", "Indexer", "MoE", "TileLang")):
        return "compute"
    return "memory"


def _short_kernel(name: str, limit: int = 72) -> str:
    name = re.sub(r"\s+", " ", name.strip())
    if len(name) <= limit:
        return name
    return name[: limit - 3] + "..."


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        import subprocess

        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "-q", "openpyxl"],
            stdout=subprocess.DEVNULL,
        )


def _event_self_time_us(evt: Any) -> float:
    for attr in ("self_cuda_time_total", "self_device_time_total", "self_cpu_time_total"):
        val = getattr(evt, attr, None)
        if val:
            return float(val)
    return 0.0


def _event_total_time_us(evt: Any) -> float:
    for attr in ("cuda_time_total", "device_time_total", "cpu_time_total"):
        val = getattr(evt, attr, None)
        if val:
            return float(val)
    return 0.0


def _events_from_key_averages(key_averages: Any) -> list[OpStat]:
    events: list[OpStat] = []
    for evt in key_averages:
        name = evt.key
        shapes = ""
        if hasattr(evt, "input_shapes") and evt.input_shapes:
            shapes = str(evt.input_shapes)
        events.append(
            OpStat(
                category=categorize_operator(name),
                name=name,
                count=int(evt.count),
                self_cuda_us=_event_self_time_us(evt),
                cuda_us=_event_total_time_us(evt),
                input_shapes=shapes,
            )
        )
    return events


def write_operator_breakdown_xlsx(
    events: list[OpStat],
    out_path: str,
    *,
    prof_start: int,
    prof_end: int,
    title: str,
    source_note: str,
) -> None:
    _ensure_openpyxl()
    from openpyxl import Workbook
    from openpyxl.styles import Font

    nsteps = max(1, prof_end - prof_start + 1)
    total_self_ms = sum(e.self_cuda_ms for e in events) / nsteps

    by_cat: dict[str, dict[str, float | int]] = defaultdict(
        lambda: {"count": 0, "self_ms": 0.0}
    )
    for evt in events:
        bucket = by_cat[evt.category]
        bucket["count"] += evt.count
        bucket["self_ms"] += evt.self_cuda_ms

    wb = Workbook()

    # Sheet 1 — category summary
    ws1 = wb.active
    ws1.title = "汇总"
    ws1.append([title])
    ws1.append([source_note])
    ws1.append([])
    ws1.append(["算子类别", "算子数/step", "ms/step", "占比 %"])
    for cat in sorted(by_cat, key=lambda c: -by_cat[c]["self_ms"]):
        count_per_step = by_cat[cat]["count"] / nsteps
        ms_per_step = by_cat[cat]["self_ms"] / nsteps
        share = (ms_per_step / total_self_ms * 100.0) if total_self_ms > 0 else 0.0
        ws1.append([cat, round(count_per_step, 1), round(ms_per_step, 2), round(share, 1)])
    ws1.append([])
    ws1.append(["GPU self-time 合计 ms/step", round(total_self_ms, 2)])
    ws1["A1"].font = Font(bold=True)

    # Sheet 2 — per-kernel by category
    ws2 = wb.create_sheet("算子明细")
    ws2.append(["算子类别", "GPU kernel / op", "数/step", "ms/step", "占比 %"])
    ranked = sorted(events, key=lambda e: -e.self_cuda_ms)
    for evt in ranked:
        if evt.self_cuda_ms <= 0:
            continue
        ms_step = evt.self_cuda_ms / nsteps
        share = (ms_step / total_self_ms * 100.0) if total_self_ms > 0 else 0.0
        ws2.append(
            [
                evt.category,
                _short_kernel(evt.name, 96),
                round(evt.count / nsteps, 1),
                round(ms_step, 2),
                round(share, 1),
            ]
        )

    # Sheet 3 — notes
    ws3 = wb.create_sheet("说明")
    notes = [
        "单位: 数/step = 每步 op 调用次数; ms/step = 每步 torch.profiler Self CUDA time 合计。",
        f"Profile window: Megatron train steps {prof_start}-{prof_end} ({nsteps} steps), rank 0 only.",
        "Self CUDA 为多 stream 上的 kernel self-time 之和，可大于 wall step time（计算/通信重叠）。",
        "算子类别由 kernel/op 名称启发式归类（GEMM / Sparse MLA / MHC / MoE / comm / copy 等）。",
        "DSV4 4-layer 默认: TileLang Sparse MLA + TileLang MHC + Megatron MoE EP (NCCL alltoall)。",
    ]
    for line in notes:
        ws3.append([line])

    # Sheet 4 — top kernels (qwen3 单算子明细 style)
    ws4 = wb.create_sheet("单算子明细")
    ws4.append(
        [
            "#",
            "kernel / op",
            "shape",
            "calls/step",
            "实测 ms/step",
            "类型",
            "占比 %",
            "说明",
        ]
    )
    ws4.append([source_note])
    top = [e for e in ranked if e.self_cuda_ms > 0][:80]
    for idx, evt in enumerate(top, start=1):
        ms_step = evt.self_cuda_ms / nsteps
        share = (ms_step / total_self_ms * 100.0) if total_self_ms > 0 else 0.0
        ws4.append(
            [
                idx,
                _short_kernel(evt.name, 80),
                evt.input_shapes or "",
                round(evt.count / nsteps, 1),
                round(ms_step, 2),
                kernel_type(evt.category, evt.name),
                round(share, 1),
                evt.category,
            ]
        )

    os.makedirs(os.path.dirname(os.path.abspath(out_path)) or ".", exist_ok=True)
    wb.save(out_path)


def export_profile_artifacts(
    prof: Any,
    *,
    prof_start: int,
    prof_end: int,
    output_txt: str,
    output_xlsx: str,
    title: str,
    source_note: str,
    trace_path: str = "",
) -> None:
    key_averages = prof.key_averages()
    sort_key = "self_device_time_total"
    sample = key_averages[0] if key_averages else None
    if sample is not None and not hasattr(sample, sort_key):
        sort_key = "self_cuda_time_total" if hasattr(sample, "self_cuda_time_total") else "self_cpu_time_total"
    table = key_averages.table(sort_by=sort_key, row_limit=200)
    os.makedirs(os.path.dirname(os.path.abspath(output_txt)) or ".", exist_ok=True)
    with open(output_txt, "w", encoding="utf-8") as f:
        f.write(f"{title}\n")
        f.write(f"{source_note}\n\n")
        f.write(table)

    events = _events_from_key_averages(key_averages)
    json_path = output_txt.replace(".txt", ".json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(
            [
                {
                    "category": e.category,
                    "name": e.name,
                    "count": e.count,
                    "self_cuda_us": e.self_cuda_us,
                    "cuda_us": e.cuda_us,
                    "input_shapes": e.input_shapes,
                }
                for e in events
            ],
            f,
            indent=2,
        )

    write_operator_breakdown_xlsx(
        events,
        output_xlsx,
        prof_start=prof_start,
        prof_end=prof_end,
        title=title,
        source_note=source_note,
    )

    if trace_path:
        prof.export_chrome_trace(trace_path)

    print(f"[PROFILER] Wrote {output_txt}", flush=True)
    print(f"[PROFILER] Wrote {json_path}", flush=True)
    print(f"[PROFILER] Wrote {output_xlsx}", flush=True)
    if trace_path:
        print(f"[PROFILER] Wrote chrome trace {trace_path}", flush=True)


def install_dsv4_profiler() -> None:
    """Patch Megatron ``train_step`` when ``LUMEN_PROF_START`` is set."""
    prof_start_raw = os.environ.get("LUMEN_PROF_START")
    if not prof_start_raw:
        return

    import functools

    import torch
    import torch.distributed as dist
    import megatron.training.training as megatron_training

    prof_start = int(prof_start_raw)
    prof_end = int(os.environ.get("LUMEN_PROF_END", str(prof_start + 2)))
    results_dir = _default_results_dir()
    output_txt = os.environ.get(
        "LUMEN_PROF_OUTPUT",
        os.path.join(results_dir, "dsv4_4layer_profile.txt"),
    )
    output_xlsx = os.environ.get(
        "LUMEN_PROF_XLSX",
        os.path.join(results_dir, "dsv4_4layer_operator_breakdown.xlsx"),
    )
    trace_path = os.environ.get("LUMEN_PROF_TRACE", "")
    record_shapes = os.environ.get("LUMEN_PROF_SHAPES", "0") == "1"
    stop_after = int(os.environ.get("LUMEN_PROF_STOP_AFTER", str(prof_end)))

    title = os.environ.get(
        "LUMEN_PROF_TITLE",
        "DSV4 Flash 4-layer Megatron pretrain — 单算子 Self CUDA 耗时",
    )
    source_note = os.environ.get(
        "LUMEN_PROF_SOURCE",
        (
            f"torch.profiler steps {prof_start}-{prof_end} | "
            f"MLA={os.environ.get('V4_SPARSE_MLA_BACKEND', '?')} "
            f"MHC={os.environ.get('MHC_BACKEND', '?')} "
            f"seq={os.environ.get('SEQ_LEN', '?')} "
            f"GBS={os.environ.get('GBS', '?')}"
        ),
    )

    state: dict[str, Any] = {"step": 0, "prof": None}
    orig_train_step = megatron_training.train_step

    @functools.wraps(orig_train_step)
    def profiled_train_step(*args, **kwargs):
        state["step"] += 1
        step = state["step"]
        rank = dist.get_rank() if dist.is_initialized() else 0

        if step == prof_start and rank == 0:
            state["prof"] = torch.profiler.profile(
                activities=[
                    torch.profiler.ProfilerActivity.CPU,
                    torch.profiler.ProfilerActivity.CUDA,
                ],
                record_shapes=record_shapes,
            )
            state["prof"].__enter__()
            print(
                f"[PROFILER] Started at step {step} (window {prof_start}-{prof_end})",
                flush=True,
            )

        result = orig_train_step(*args, **kwargs)

        if step == prof_end and rank == 0 and state["prof"] is not None:
            state["prof"].__exit__(None, None, None)
            export_profile_artifacts(
                state["prof"],
                prof_start=prof_start,
                prof_end=prof_end,
                output_txt=output_txt,
                output_xlsx=output_xlsx,
                title=title,
                source_note=source_note,
                trace_path=trace_path,
            )
            state["prof"] = None

        if stop_after and step >= stop_after:
            if rank == 0:
                print(f"[PROFILER] Stopping after step {step}", flush=True)
            raise SystemExit(0)

        return result

    megatron_training.train_step = profiled_train_step
    print(
        f"[PROFILER] Armed steps {prof_start}-{prof_end}, "
        f"txt={output_txt}, xlsx={output_xlsx}",
        flush=True,
    )


def main(argv: list[str] | None = None) -> int:
    """Convert a saved profile JSON to xlsx (offline)."""
    import argparse

    parser = argparse.ArgumentParser(description="Export DSV4 profile JSON to operator xlsx")
    parser.add_argument("json_path", help="Profile JSON from dsv4_profiler")
    parser.add_argument(
        "-o",
        "--output",
        default="",
        help="Output xlsx path (default: same dir, .xlsx suffix)",
    )
    parser.add_argument("--prof-start", type=int, default=3)
    parser.add_argument("--prof-end", type=int, default=5)
    parser.add_argument("--title", default="DSV4 Flash 4-layer — operator breakdown")
    parser.add_argument("--source", default="offline export from profile JSON")
    args = parser.parse_args(argv)

    with open(args.json_path, encoding="utf-8") as f:
        raw = json.load(f)
    events = [
        OpStat(
            category=item["category"],
            name=item["name"],
            count=int(item["count"]),
            self_cuda_us=float(item["self_cuda_us"]),
            cuda_us=float(item.get("cuda_us", 0.0)),
            input_shapes=item.get("input_shapes", ""),
        )
        for item in raw
    ]
    out = args.output or args.json_path.replace(".json", "_operator_breakdown.xlsx")
    write_operator_breakdown_xlsx(
        events,
        out,
        prof_start=args.prof_start,
        prof_end=args.prof_end,
        title=args.title,
        source_note=args.source,
    )
    print(f"Wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
